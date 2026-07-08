from datetime import datetime
from decimal import Decimal

import mysql.connector
from openpyxl import load_workbook

from scripts.export_contracts_excel import COLUMNS, export_contracts


class FakeCursor:
    def __init__(self, rows):
        self.rows = rows
        self.query = None

    def execute(self, query):
        self.query = query

    def fetchall(self):
        return self.rows

    def close(self):
        pass


class FakeConnection:
    def __init__(self, rows):
        self.cursor_instance = FakeCursor(rows)

    def cursor(self, dictionary=True):
        return self.cursor_instance

    def is_connected(self):
        return True

    def close(self):
        pass


def test_export_contracts_excel_exports_all_contracts(monkeypatch, tmp_path):
    rows = [
        {
            "COD_INSTITUICAO": Decimal("123"),
            "COD_COMPARTILHADO": None,
            "NOME_INSTITUICAO": "Cliente A",
            "SERVICOS_CONTRATADOS": "Individual, API",
            "NUM_CONTRATO": "C-001",
            "VALOR_EXCEDENTE": Decimal("1.25"),
            "DT_CORTE_INICIAL": datetime(2026, 1, 1),
            "DT_FIM": datetime(2026, 12, 31),
            "FREQUENCIA_CORTE": "Anual",
            "NUM_AC_CONTRATADOS": None,
            "FL_ACESSOS_ILIMITADOS": 1,
        },
        {
            "COD_INSTITUICAO": Decimal("456"),
            "COD_COMPARTILHADO": Decimal("789"),
            "NOME_INSTITUICAO": "Cliente B",
            "SERVICOS_CONTRATADOS": "Lote",
            "NUM_CONTRATO": "C-002",
            "VALOR_EXCEDENTE": None,
            "DT_CORTE_INICIAL": datetime(2026, 2, 1),
            "DT_FIM": datetime(2026, 2, 28),
            "FREQUENCIA_CORTE": "Mensal",
            "NUM_AC_CONTRATADOS": Decimal("5000"),
            "FL_ACESSOS_ILIMITADOS": 0,
        },
    ]
    connection = FakeConnection(rows)
    monkeypatch.setattr(mysql.connector, "connect", lambda **kwargs: connection)

    output_path = tmp_path / "contratos.xlsx"
    total = export_contracts(str(output_path))

    workbook = load_workbook(output_path)
    worksheet = workbook["Planilha1"]

    assert total == 2
    assert [worksheet.cell(1, column).value for column in range(1, 12)] == COLUMNS
    assert worksheet.auto_filter.ref == "A1:M3"
    assert "FL_MONITORAR_CONTRATO" not in connection.cursor_instance.query
    assert "WHERE" not in connection.cursor_instance.query.upper()

    assert worksheet["A2"].value == 123
    assert worksheet["B2"].value is None
    assert worksheet["F2"].value is None
    assert worksheet["G2"].value == 1.25
    assert worksheet["H2"].value == datetime(2026, 1, 1)
    assert worksheet["I2"].value == datetime(2026, 12, 31)
    assert worksheet["K2"].value == "ILIMITADO"

    assert worksheet["A3"].value == 456
    assert worksheet["B3"].value == 789
    assert worksheet["F3"].value is None
    assert worksheet["G3"].value is None
    assert worksheet["K3"].value == 5000
