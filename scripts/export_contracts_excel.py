import argparse
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import mysql.connector
from mysql.connector import Error
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.config import Config


SHEET_NAME = "Planilha1"
COLUMNS = [
    "Codigo Instituicao",
    "Cod Compartilhado",
    "Nome Instituicao",
    "Serviços Contratados",
    "Numero Contrato",
    "Prazo Contrato meses",
    "Valor Excedente",
    "data de corte início",
    "data de corte final",
    "Frequencia",
    "acessos contratados",
]
COLUMN_WIDTHS = {
    "A": 20.13,
    "B": 20.67,
    "C": 42.25,
    "D": 21.03,
    "E": 17.41,
    "F": 21.66,
    "G": 18.22,
    "H": 19.01,
    "I": 18.34,
    "J": 13.15,
    "K": 21.26,
}
CURRENCY_FORMAT = '_-"R$ "* #,##0.00_-;"-R$ "* #,##0.00_-;_-"R$ "* \\-??_-;_-@_-'


def _scalar(value: Any) -> Any:
    if isinstance(value, Decimal):
        return int(value) if value == value.to_integral_value() else float(value)
    return value


def _accesses(row: dict[str, Any]) -> Any:
    if int(row.get("FL_ACESSOS_ILIMITADOS") or 0) == 1:
        return "ILIMITADO"
    return _scalar(row.get("NUM_AC_CONTRATADOS"))


def _fetch_contracts() -> list[dict[str, Any]]:
    query = """
        SELECT
            c.COD_INSTITUICAO,
            c.COD_COMPARTILHADO,
            i.NOME_INSTITUICAO,
            c.SERVICOS_CONTRATADOS,
            i.NUM_CONTRATO,
            c.VALOR_EXCEDENTE,
            c.DT_CORTE_INICIAL,
            i.DT_FIM,
            c.FREQUENCIA_CORTE,
            c.NUM_AC_CONTRATADOS,
            c.FL_ACESSOS_ILIMITADOS
        FROM TB_CONTRATO c
        JOIN TB_INSTITUICAO i
          ON i.COD_INSTITUICAO = c.COD_INSTITUICAO
        ORDER BY c.COD_INSTITUICAO, i.NUM_CONTRATO, c.SERVICOS_CONTRATADOS
    """
    conn = None
    cursor = None
    try:
        conn = mysql.connector.connect(
            host=Config.DB_HOST,
            user=Config.DB_USER,
            password=Config.DB_PASS,
            database=Config.DB_NAME,
        )
        cursor = conn.cursor(dictionary=True)
        cursor.execute(query)
        return cursor.fetchall()
    except Error as exc:
        raise RuntimeError(f"Erro ao consultar contratos no MySQL: {exc}") from exc
    finally:
        if cursor:
            cursor.close()
        if conn and conn.is_connected():
            conn.close()


def _row_values(row: dict[str, Any]) -> list[Any]:
    return [
        _scalar(row.get("COD_INSTITUICAO")),
        _scalar(row.get("COD_COMPARTILHADO")),
        row.get("NOME_INSTITUICAO"),
        row.get("SERVICOS_CONTRATADOS"),
        row.get("NUM_CONTRATO"),
        None,
        _scalar(row.get("VALOR_EXCEDENTE")),
        row.get("DT_CORTE_INICIAL"),
        row.get("DT_FIM"),
        row.get("FREQUENCIA_CORTE"),
        _accesses(row),
    ]


def export_contracts(output_path: str) -> int:
    rows = _fetch_contracts()
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME
    worksheet.sheet_view.zoomScale = 90

    header_font = Font(name="Arial", size=11, bold=True)
    body_font = Font(name="Arial", size=11)
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")
    center = Alignment(horizontal="center", vertical="center")

    for column_number, header in enumerate(COLUMNS, start=1):
        cell = worksheet.cell(row=1, column=column_number, value=header)
        cell.font = header_font
        cell.fill = yellow_fill
        cell.alignment = center

    for row_number, row in enumerate(rows, start=2):
        worksheet.row_dimensions[row_number].height = 14.25
        for column_number, value in enumerate(_row_values(row), start=1):
            cell = worksheet.cell(row=row_number, column=column_number, value=value)
            cell.font = body_font
            cell.alignment = center
            if column_number in {8, 9} and isinstance(value, datetime | date):
                cell.number_format = "d/m/yyyy"
            elif column_number == 7 and value is not None:
                cell.number_format = CURRENCY_FORMAT
            elif column_number == 11 and isinstance(value, int):
                cell.number_format = "#,##0"

    worksheet.row_dimensions[1].height = 14.25
    for column_letter, width in COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column_letter].width = width

    last_row = max(worksheet.max_row, 1)
    worksheet.auto_filter.ref = f"A1:M{last_row}"
    workbook.save(output)
    return len(rows)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Exporta a TB_CONTRATO para XLSX no layout da planilha original."
    )
    parser.add_argument(
        "--output",
        default="/tmp/contratos_exportados.xlsx",
        help="Arquivo XLSX de saida.",
    )
    args = parser.parse_args()

    total = export_contracts(args.output)
    print(f"XLSX gerado: {args.output}")
    print(f"Contratos exportados: {total}")


if __name__ == "__main__":
    main()
